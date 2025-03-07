import sys
import threading
from tkinter import filedialog, messagebox

from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from Add_picture import input_L2word, input_L3word, input_L5word, input_S2word, input_HPC64word, input_L6word


def open_L3file():
    L3_filepath = filedialog.askopenfilename(filetypes=[("Word files", ".docx")])
    if L3_filepath:
        thread = threading.Thread(target=input_L3word, args=(L3_filepath,), daemon=True)
        thread.start()
        print("文档插入图片完成")
        task_completed()
    else:
        print("No file selected.")


def load_excel():
    file_path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if file_path:
        wb = load_workbook(filename=file_path)
        ws = wb.active
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append(row)

        # 保存到另一个Excel文件
        new_wb = Workbook()
        new_ws = new_wb.active
        for row in data:
            new_ws.append(row)
        new_wb.save("L3巡检密码.xlsx")
        print("数据已保存到output.xlsx")
        task_completed()


def open_L2file():
    L2_filepath = filedialog.askopenfilename(filetypes=[("Word files", ".docx")])
    if L2_filepath:
        thread = threading.Thread(target=input_L2word, args=(L2_filepath,), daemon=True)
        thread.start()
        print("文档插入图片完成")
        task_completed()
    else:
        print("No file selected.")


def open_L5file():
    L5_filepath = filedialog.askopenfilename(filetypes=[("Word files", ".docx")])
    if L5_filepath:
        thread = threading.Thread(target=input_L5word, args=(L5_filepath,), daemon=True)
        thread.start()
        print("文档插入图片完成")
        task_completed()
    else:
        print("No file selected.")


def open_S2file():
    S2_filepath = filedialog.askopenfilename(filetypes=[("Word files", ".docx")])
    if S2_filepath:
        thread = threading.Thread(target=input_S2word, args=(S2_filepath,), daemon=True)
        thread.start()
        print("文档插入图片完成")
        task_completed()
    else:
        print("No file selected.")


def open_HPC64file():
    HPC64_filepath = filedialog.askopenfilename(filetypes=[("Word files", ".docx")])
    if HPC64_filepath:
        thread = threading.Thread(target=input_HPC64word, args=(HPC64_filepath,), daemon=True)
        thread.start()
        print("文档插入图片完成")
        task_completed()
    else:
        print("No file selected.")


def open_L6ZZJD_file():
    L6_filepath = filedialog.askopenfilename(filetypes=[("Word files", ".docx")])
    if L6_filepath:
        thread = threading.Thread(target=input_L6word, args=(L6_filepath,), daemon=True)
        thread.start()
        print("文档插入图片完成")
        task_completed()
    else:
        print("No file selected.")


def run_L3py():
    from OPEN_ALL_GO import L3py_files
    threading.Thread(target=L3py_files, daemon=True).start()

def run_L3GLQFBD():
    from OPEN_ALL_GO import L3GLQoy_run
    threading.Thread(target=L3GLQoy_run, daemon=True).start()


def run_L2py():
    from OPEN_ALL_GO import L2py_files
    L2_thread = threading.Thread(target=L2py_files, daemon=True)
    L2_thread.start()



def run_L5py():
    from OPEN_ALL_GO import L5py_files
    threading.Thread(target=L5py_files, daemon=True).start()



def run_S2py():
    from OPEN_ALL_GO import S2py_files
    threading.Thread(target=S2py_files, daemon=True).start()



def run_HPC64py():
    from OPEN_ALL_GO import HPC64py_files
    HPC64py_thread = threading.Thread(target=HPC64py_files, daemon=True)
    HPC64py_thread.start()



def run_L6ZZJDpy():
    from OPEN_ALL_GO import L6ZZJD_files
    L6ZZJDpy_thread = threading.Thread(target=L6ZZJD_files, daemon=True)
    L6ZZJDpy_thread.start()



def task_completed():
    messagebox.showinfo("任务完成", "任务已经成功完成！")


if __name__ == "__main__":
    print('')
